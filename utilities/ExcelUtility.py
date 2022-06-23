import openpyxl
# from openpyxl.styles import PatternFill

#for row count:
def row_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

#for column count
def column_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

#for reading data
def read_data(file,sheetName,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.cell(row,column).value)

#for writing data
def write_data(file,sheetName,row,column,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row,column).value = data
    workbook.save(file)

#for filling green colour
# def fillGreenColour(file,sheetName,row,column):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     greenFill=PatternFill(start_color='60b212',
#                           end_color='60b212',
#                           fill_type='solid')
#     sheet.cell(row,column).fill=greenFill
#     workbook.save(file)

#for filling red colour
# def fillRedColour(file,sheetName,row,column):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     redFill=PatternFill(start_color='ff0000',
#                           end_color='ff0000',
#                           fill_type='solid')
#     sheet.cell(row,column).fill=redFill
#     workbook.save(file)